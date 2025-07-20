from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect
from django.template import loader
from django.http import HttpResponse
from django import template
import datetime
from dateutil.relativedelta import *

from django.templatetags.static import static

import calendar
from django.forms import inlineformset_factory
from django.views.generic import ListView, CreateView, DetailView, DeleteView, UpdateView
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils.decorators import method_decorator
from dateutil.rrule import rrule, MONTHLY, DAILY, WEEKLY
from django.db.models import Sum
from django.db.models import Q
import io
from openpyxl import Workbook
#from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
import xlwt





#from django.contrib.auth.mixins import LoginRequiredMixin


#orangepi comment


#------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------

# Felhasználó kezelés moduljai

#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm
 
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.contrib.auth import update_session_auth_hash
from django.utils.translation import gettext as _

#------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------

# Saját file-ok importjai

#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------

from app.forms import *
from app.models import *
from app.filters import *
from django.contrib.auth.models import User
from app.decorators import unauthenticated_user, allowed_users, vezetoi, emusok
#from app.apps import TervezettHaviBevetel, TervezettHaviOrak, ElozoHaviBevetel, LemondottOrakSzama, OsszesOraKategoriak, JelenletiGen, JelenletiReGen

#------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------

# Függvények:



def GyerekKodGeneralo():
    kod=""
    gyerekek = Gyerek.objects.all()
    for gyerek in gyerekek:
        if not gyerek.kod :
            kod = gyerek.nev[:2]+gyerek.nev.split(" ")[1]
            gyerek.kod = kod
            gyerek.save()


def TervezettHaviBevetel():
    ev = datetime.datetime.now().year
    honap = datetime.datetime.now().month
    ehavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month)
    ehavi_egyedi=Egyedi.objects.filter(idopont__year=datetime.datetime.now().year, idopont__month=datetime.datetime.now().month)
    ora_bevetel = 0
    egyedi_bevetel = 0
    for ora in ehavi_orak:
        csoport = ora.csoport
        ar = ora.csoport.foglalkozas_tipus.ar
        gyszam = GyerektoCsoport.objects.filter(csoport=csoport).count()
        ora_bevetel = ora_bevetel + ar * gyszam
    for egyedi in ehavi_egyedi:
        egyedi_bevetel =  egyedi_bevetel + egyedi.ar * egyedi.letszam
    return (ora_bevetel, egyedi_bevetel)
    #befejezve

def TervezettHaviOrak():
    ehavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month)
    ehavi_orak_szama=ehavi_orak.count()
    return (ehavi_orak_szama)
    #befejezve

def ElozoHaviBevetel():
    honap = datetime.datetime.now().month
    oradijak = 0
    egyedi_dijak = 0
    bevetelek = 0
    netto_bevetel = 0
    
    
    if honap == 1:
        multhavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year-1, kezdes__month=12).filter(megtartott=True)
        multhavi_egyediek=Egyedi.objects.filter(idopont__year=datetime.datetime.now().year-1, idopont__month=12).filter(megtartva=True)
        multhavi_befizetesek=Befizetes.objects.filter(datum__year=datetime.datetime.now().year-1, datum__month=12)
    else:
        multhavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month-1).filter(megtartott=True)
        multhavi_egyediek=Egyedi.objects.filter(idopont__year=datetime.datetime.now().year, idopont__month=datetime.datetime.now().month-1).filter(megtartva=True)
        multhavi_befizetesek=Befizetes.objects.filter(datum__year=datetime.datetime.now().year, datum__month=datetime.datetime.now().month-1)
    
    for ora in multhavi_orak:
        oraert_fizetnek=Jelenleti.objects.filter(ora=ora).filter( Q(status=1) | Q(status=2) ) # Az órához tartozó jelenléti, ahol jele voltak, vagy 100% a lemondás
        jelen_voltak=oraert_fizetnek.count() # fizető gyerekek száma az órán
        ar = ora.csoport.foglalkozas_tipus.ar # az óra ára
        oradijak = oradijak + ar * jelen_voltak # az óra ára szorozva a résztvevőkkel, kummulálva az előző órával
    
    for egyedi in multhavi_egyediek:
        egyedi_dijak = egyedi_dijak + egyedi.ar*egyedi.letszam #egyedi program ára szorozva a résztvevőkkel, kummulálva
    
    for befizetes in multhavi_befizetesek:
        bevetelek = bevetelek + befizetes.osszeg
    
    netto_bevetel = bevetelek + egyedi_dijak - oradijak
    
    return (oradijak, egyedi_dijak, bevetelek, netto_bevetel)

def LemondottOrakSzama():
    honap = datetime.datetime.now().month
    szazszazalek = 0
    nullaszazalek = 0

    if honap == 1:
        multhavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year-1, kezdes__month=12).filter(megtartott=True)
    else:
        multhavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month-1).filter(megtartott=True)

    for ora in multhavi_orak:
        szazszazalek = szazszazalek + Jelenleti.objects.filter(ora=ora).filter(status=2).count()
        nullaszazalek = nullaszazalek + Jelenleti.objects.filter(ora=ora).filter(status=3).count()

    osszesen = szazszazalek + nullaszazalek
    return (osszesen, szazszazalek, nullaszazalek)

def OsszesOraKategoriak():
    egyeni=0
    paros=0
    csoportos=0
    csterapia=0
    szuloi=0
    egyeb=0
    
    honap = datetime.datetime.now().month
    if honap == 1:
        multhavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year-1, kezdes__month=12).filter(megtartott=True)
    else:
        multhavi_orak=Ora.objects.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month-1).filter(megtartott=True)
    
    osszes = multhavi_orak.count()
    for ora in multhavi_orak:
        cat = ora.csoport.foglalkozas_tipus.cat
        if cat == 'Egyéni':
            egyeni = egyeni + 1
        elif cat == 'Páros':
            paros = paros + 1
        elif cat == 'Csoportos':
            csoportos = csoportos + 1
        elif cat == 'Család terápia':
            csterapia = csterapia + 1
        elif cat == 'Szülői':
            szuloi = szuloi + 1
        elif cat == 'Egyéb':
            egyeb = egyeb + 1
    return (egyeni, paros, csoportos, csterapia, szuloi, egyeb, osszes)
    
   


    
'''




context = {}
    eo = datetime.time(0,0,0)
    uo = datetime.time(23,59,59)
    now = datetime.datetime.now()
    now_start = datetime.datetime.combine(now, eo)
    now_end = datetime.datetime.combine(now, uo)
    egyeni_szul_konz_orak_szama=0
    megtartott_egyeni_szul_konz_orak_szama=0
    napiuzi = Napiuzi.objects.last()


    groupja = request.user.groups.all()[0].name
    guru = request.user.guru
    gurucsoportok = GurutoCsoport.objects.filter(guru=guru)
    now = datetime.datetime.now()
    utolso_nap = calendar.monthrange(int(now.strftime("%Y")), int(now.strftime("%m")))[1]
    honap_kezdete = datetime.date(int(now.strftime("%Y")), int(now.strftime("%m")), 1)
    honap_vege = datetime.date(int(now.strftime("%Y")), int(now.strftime("%m")), utolso_nap)

    #egyéni szülői konzultációs darabszám az adott hónapban
    for gurucsoport in gurucsoportok:
        tipusnev = gurucsoport.csoport.foglalkozas.tipus.nev
        if 'Szülői konz. - Egyéni' in tipusnev:
            egyeni_szul_konz_orak = gurucsoport.csoport.ora_set.filter(kezdes__range=[honap_kezdete, honap_vege])
            egyeni_szul_konz_orak_szama = egyeni_szul_konz_orak.count()  #egyéni szülői konzultációs darabszám az adott hónapban
            megtartott_egyeni_szul_konz_orak = egyeni_szul_konz_orak.filter(megtartott=True) #megtartott szülői konzultációs darabszám az adott hónapban
            megtartott_egyeni_szul_konz_orak_szama = megtartott_egyeni_szul_konz_orak.count()

    qs = GurutoCsoport.objects.values_list('csoport', flat=True).filter(guru=guru)
    mai_orak = Ora.objects.filter(csoport__in=qs).filter(kezdes__range=[now_start, now_end])
    mai_orak_szama = mai_orak.count()
    guru_orak = Ora.objects.filter(csoport__in=qs)

'''
def MissingJelenletiGen():

    orak = Ora.objects.all()
    for ora in orak:
        jelenletik = Jelenleti.objects.filter(ora=ora)
        if not jelenletik: # ha még az órához nincs jelenléti adat, létrehozza azt
            csoport = ora.csoport
            gyerekektocsoport = GyerektoCsoport.objects.filter(csoport=csoport)
            for gyerektocsoport in gyerekektocsoport:
                gyerek = gyerektocsoport.gyerek
                jelenlet = Jelenleti(gyerek=gyerek, ora=ora, status=0)
                jelenlet.save()
    return

def JelenletiGen (oraid):
    ora = Ora.objects.get(id=oraid)
    csoport = ora.csoport
    gyerekektocsoport = GyerektoCsoport.objects.filter(csoport=csoport)
    for gyerektocsoport in gyerekektocsoport:
                gyerek = gyerektocsoport.gyerek
                jelenlet = Jelenleti(gyerek=gyerek, ora=ora, status=0)
                jelenlet.save()

def JelenletiReGen (oraid):
    ora = Ora.objects.get(id=oraid)
    if not ora.megtartott or ora.lezart:
        Jelenleti.objects.filter(ora=ora).delete()
        csoport = ora.csoport
        gyerekektocsoport = GyerektoCsoport.objects.filter(csoport=csoport)
        for gyerektocsoport in gyerekektocsoport:
                    gyerek = gyerektocsoport.gyerek
                    jelenlet = Jelenleti(gyerek=gyerek, ora=ora, status=0)
                    jelenlet.save()
    




# Views :

#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
@login_required(login_url="/login/")
@vezetoi
def index(request):
    context = {}
    osszes_befizetes = 0
    havi_orak_szama = 0
    ora_ar = 0
    kintlevoseg = 0
    ar = 0
    megtartott_orak_szama = 0
    
    gyerekek_szama=Gyerek.objects.filter(aktiv=True).count
    havi_orak = Ora.objects.all().filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month)
    havi_orak_szama = havi_orak.count()
    megtartott_orak_szama = havi_orak.filter(megtartott=True).count()
    

    context['havi_orak_szama']=havi_orak_szama
    context['megtartott_orak_szama']=megtartott_orak_szama
    context['osszes_befizetes']=osszes_befizetes
    context['gyerekek_szama']=gyerekek_szama

    return render(request, "index.html", context)

@login_required(login_url="/login/")
@vezetoi
def AttekintoView(request):
    context = {}
    tervezett_havi_bevetel=TervezettHaviBevetel()
    tervezett_havi_orak = TervezettHaviOrak()
    elozo_havi_bevetel = ElozoHaviBevetel()
    lemondott_orak = LemondottOrakSzama()
    kategoriak = OsszesOraKategoriak()
    aktiv_terapiak_szama=Csoport.objects.filter(aktiv=True).count
   
    
    context['tervezett_havi_bevetel']=tervezett_havi_bevetel
    context['tervezett_havi_orak']=tervezett_havi_orak
    context['elozo_havi_bevetel']=elozo_havi_bevetel
    context['lemondott_orak']=lemondott_orak
    context['kategoriak']=kategoriak
    context['aktiv_terapiak_szama']=aktiv_terapiak_szama
    


    return render(request, "attekinto.html", context)




@login_required(login_url="/login/")
def guru_lap(request):
    context = {}
    eo = datetime.time(0,0,0)
    uo = datetime.time(23,59,59)
    now = datetime.datetime.now()
    now_start = datetime.datetime.combine(now, eo)
    now_end = datetime.datetime.combine(now, uo)
    egyeni_szul_konz_orak_szama=0
    megtartott_egyeni_szul_konz_orak_szama=0
    napiuzi = Napiuzi.objects.last()
    
    guru = request.user.guru
    terapeuta_nev=guru.nev
    qs = GurutoCsoport.objects.values_list('csoport', flat=True).filter(guru=guru)
    csoportok = Csoport.objects.filter(id__in=qs).filter(aktiv=True)
   
    utolso_nap = calendar.monthrange(int(now.strftime("%Y")), int(now.strftime("%m")))[1]
    honap_kezdete = datetime.date(int(now.strftime("%Y")), int(now.strftime("%m")), 1)
    honap_vege = datetime.date(int(now.strftime("%Y")), int(now.strftime("%m")), utolso_nap)

    #egyéni szülői konzultációs darabszám az adott hónapban
    for csoport in csoportok:
        tipusnev = csoport.foglalkozas_tipus.nev
        if 'Szülői konz. - Egyéni' in tipusnev:
            egyeni_szul_konz_orak = csoport.ora_set.filter(kezdes__range=[honap_kezdete, honap_vege])
            egyeni_szul_konz_orak_szama = egyeni_szul_konz_orak.count() + egyeni_szul_konz_orak_szama  #egyéni szülői konzultációs darabszám az adott hónapban
            megtartott_egyeni_szul_konz_orak = egyeni_szul_konz_orak.filter(megtartott=True) #megtartott szülői konzultációs darabszám az adott hónapban
            megtartott_egyeni_szul_konz_orak_szama = megtartott_egyeni_szul_konz_orak.count() + megtartott_egyeni_szul_konz_orak_szama

    
    mai_orak = Ora.objects.filter(csoport__in=csoportok).filter(kezdes__range=[now_start, now_end])
    mai_orak_szama = mai_orak.count()
    guru_orak = Ora.objects.filter(csoport__in=csoportok)
                                                                       

    #aznapi óraszám

    #üzenet
    context = {'guru':guru, 'egyeni_szul_konz_orak_szama':egyeni_szul_konz_orak_szama, 'megtartott_egyeni_szul_konz_orak_szama':megtartott_egyeni_szul_konz_orak_szama}
    context['mai_orak_szama']=mai_orak_szama
    context['guru_orak']=guru_orak
    context['napiuzi']=napiuzi
    context['terapeuta_nev']=terapeuta_nev
    
    return render (request, 'index_guru.html', context)


@login_required(login_url="/login/")
def szulo_lap(request):
    szulo = guru = request.user.szulo
    #guru = Guru.objects.get(id=pk)
    return render (request, 'index_szulo.html')


@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:
        
        load_template = request.path.split('/')[-1]
        html_template = loader.get_template( load_template )
        return HttpResponse(html_template.render(context, request))
        
    except template.TemplateDoesNotExist:

        html_template = loader.get_template( 'page-404.html' )
        return HttpResponse(html_template.render(context, request))

    except:
    
        html_template = loader.get_template( 'page-500.html' )
        return HttpResponse(html_template.render(context, request))



#------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------

# Felhasználó kezelés view-jai



#-------------------------------------------Gyerek views--------------------------------------------------------------------------------------------
@login_required(login_url='login')
@vezetoi
def GyerekCreate(request):
    form = GyerekForm()
    gyerek = Gyerek.objects.none()
    

    #GyerekInlineFormSet = inlineformset_factory(Gyerek, Csalad, fields=('szulo',), extra=2, max_num=2)
    formset = CsaladFormSet()
    
    context ={}
    lekerdezes = 0

    if request.method == 'POST':
        form = GyerekForm(request.POST)
        #formset = GyerekInlineFormSet(request.POST)
        formset = CsaladFormSet(request.POST)

        if form.is_valid():
            created_gyerek = form.save(commit=False)
            nev = created_gyerek.nev
            kod = nev[:2]+nev.split(" ")[1]
            created_gyerek.kod = kod
            
            #formset = GyerekInlineFormSet(request.POST, instance=created_gyerek)
            formset = CsaladFormSet(request.POST, instance=created_gyerek)
            if formset.is_valid():
                created_gyerek.save()
                formset.save()
                return redirect('gyerek_lista')

    context['lekerdezes'] = lekerdezes
    context['Form_neve'] = 'Gyerek felvétel űrlap'
    context = {
        "form": form,
        "formset": formset,
        "gyerek": gyerek,
        }
    
    return render(request, 'gyerek.html', context)


@login_required(login_url="/login/")
@vezetoi
def GyerekList(request):
    context = {}
    gyerekek = Gyerek.objects.all()
    #myFilter = GyerekFilter(request.GET, queryset=gyerekek)
    #gyerekek = myFilter.qs

    #paginator = Paginator(gyerekek, 30) # Show 30 contacts per page.
    #page_number = request.GET.get('page')
    #gyerekek = paginator.get_page(page_number)
    
    context['gyerekek'] = gyerekek
    #context['myFilter'] = myFilter
    return render(request, 'gyerek_list.html', context)


@login_required(login_url="/login/")
@vezetoi
def GyerekListExport(request):
    
    gyerekek = GyerekFilter(request.GET, queryset=Gyerek.objects.all()).qs
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="gyerekek.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Gyerekek')
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['id', 'Név', 'Születési idő', 'Anyja neve', 'Aktív']

    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style.font.bold = False
    row_num = 0

    for item in gyerekek:
        row_num += 1
        ws.write(row_num, 0, str(getattr(item, 'id')), font_style)
        ws.write(row_num, 1, str(getattr(item, 'nev')), font_style)
        ws.write(row_num, 2, str(getattr(item, 'szul_ido')), font_style)
        ws.write(row_num, 3, str(getattr(item, 'anyja_neve')), font_style)
        ws.write(row_num, 4, str(getattr(item, 'aktiv')), font_style)
        
    
    wb.save(response)
    return response

@login_required(login_url="/login/")
@vezetoi
def GyerekUpdate(request, pk):
    gyerek = Gyerek.objects.get(id=pk)
    form = GyerekUpdateForm(instance=gyerek)
    
    formset = CsaladFormSet(instance=gyerek)
    context ={}
    lekerdezes = 1
    context['lekerdezes'] = lekerdezes
    gyerektocsoportok = GyerektoCsoport.objects.filter(gyerek=gyerek)
    orak = Ora.objects.none()
    
    for gyerektocsoport in gyerektocsoportok:
        csoport = gyerektocsoport.csoport
        orak |= Ora.objects.filter(csoport=csoport)
    
    if request.method == 'POST':
        form = GyerekUpdateForm(request.POST, instance=gyerek)
        
        if form.is_valid():
            created_gyerek = form.save(commit=False)
            #nev = created_gyerek.nev
            #kod = nev[:2]+nev.split(" ")[1]
            #created_gyerek.kod = kod
            
            formset = CsaladFormSet(request.POST, instance=created_gyerek)
            if formset.is_valid():
                created_gyerek.save()
                formset.save()
                return redirect('gyerek_lista')
    
    context = {'gyerek':gyerek, 'gyerektocsoportok':gyerektocsoportok,'form':form, 'orak':orak, 'formset':formset}
    return render(request, 'gyerek.html', context)



#-------------------------------------------Szülő views--------------------------------------------------------------------------------------------
@login_required(login_url='login')
@vezetoi
def SzuloCreate(request):
    form = SzuloForm()
    context ={}
    
    if request.method == 'POST':
        form = SzuloForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('gyerek_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Szűlő felvétel űrlap'
    return render(request, 'create_form.html', context)



@login_required(login_url="/login/")
@vezetoi
def SzuloList(request):
    context = {}
    szulok = Szulo.objects.all()
    context['szulok'] = szulok
    return render(request, 'szulo_list.html', context)

@login_required(login_url="/login/")
@vezetoi
def SzuloListExport(request):
    
    szulok = SzuloFilter(request.GET, queryset=Szulo.objects.all()).qs
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="szülök.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Szülők')
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['id', 'Név', 'Telefon', 'E-mail', 'Levelezési cím', 'Szerződő', 'Aktív']

    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style.font.bold = False
    row_num = 0

    for item in szulok:
        row_num += 1
        ws.write(row_num, 0, str(getattr(item, 'id')), font_style)
        ws.write(row_num, 1, str(getattr(item, 'nev')), font_style)
        ws.write(row_num, 2, str(getattr(item, 'telefon')), font_style)
        ws.write(row_num, 3, str(getattr(item, 'email')), font_style)
        ws.write(row_num, 4, str(getattr(item, 'lev_cim')), font_style)
        ws.write(row_num, 5, str(getattr(item, 'szerzodo')), font_style)
        ws.write(row_num, 6, str(getattr(item, 'aktiv')), font_style)
    
    wb.save(response)
    return response


@login_required(login_url="/login/")
@vezetoi
def SzuloUpdate(request, pk):
    
    szulo = Szulo.objects.get(id=pk)
    form = SzuloForm(instance=szulo)
    gyerekek = Csalad.objects.filter(szulo=szulo)
    
    if request.method == 'POST':
        form = SzuloForm(request.POST, instance=szulo)
        if form.is_valid():
            form.save()
            return redirect('szulo_lista')


    context = {'szulo':szulo, 'gyerekek':gyerekek, 'form':form}
    return render(request, 'szulo.html', context)

#------------------------------------------- Helyszín views--------------------------------------------------------------------------------------------
@login_required(login_url="/login/")
def HelyszinView(request):
    form = HelyszinForm()
    context ={}
    if request.method == 'POST':
        form = HelyszinForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('helyszin_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Helyszín felvétel űrlap'
    return render(request, 'create_form.html', context)


@login_required(login_url="/login/")
def HelyszinList(request):
    context = {}
    helyszinek = Helyszin.objects.all()
    context['helyszinek'] = helyszinek
    
    return render(request, 'helyszin_list.html', context)

@login_required(login_url='login')
@vezetoi
def HelyszinUpdate(request, pk):
    helyszin=Helyszin.objects.get(id=pk)
    form = HelyszinForm(instance=helyszin)
    context ={}
    if request.method == 'POST':
        form = HelyszinForm(request.POST, instance=helyszin)
        if form.is_valid():
            form.save()
            return redirect('helyszin_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Helyszín módosító űrlap'
    return render(request, 'create_form.html', context)
#-------------------------------------------Guru views--------------------------------------------------------------------------------------------

@login_required(login_url='login')
@vezetoi
def GuruCreate(request):
    form = GuruForm()
    context ={}
    if request.method == 'POST':
        form = GuruForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('guru_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Terapeuta felvétel űrlap'
    return render(request, 'create_form.html', context)

@login_required(login_url='login')
@vezetoi
def GuruList(request):
    context = {}
    guruk = Guru.objects.all()
    myFilter = GuruFilter(request.GET, queryset=guruk)
    guruk = myFilter.qs
    context['guruk'] = guruk
    context['myFilter'] = myFilter
    return render(request, 'guru_list.html', context)


@login_required(login_url='login')
# A guru saját magát updatelheti?
def GuruUpdate(request, pk):
    guru = Guru.objects.get(id=pk)
    form = GuruForm(instance=guru)
    context ={}
    terapiak = GurutoCsoport.objects.filter(guru=guru)

    if request.method == 'POST':

        form = GuruForm(request.POST, request.FILES, instance=guru)
        if form.is_valid():
            form.save()
            return redirect('guru_lista')

    context = {'form': form, 'guru':guru, 'terapiak':terapiak}
    context['Form_neve'] = 'Terapeuta módosítás űrlap'
    return render(request, 'guru.html', context)

#-------------------------------------------Foglalkozás típus views--------------------------------------------------------------------------------------------

@login_required(login_url='login')
@vezetoi
def Foglalkozas_tipusCreate(request):
    form = Foglalkozas_tipusForm()
    context ={}
    if request.method == 'POST':
        form = Foglalkozas_tipusForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('foglalkozas_tipus_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Foglalkozás típus felvétel űrlap'
    return render(request, 'create_form.html', context)

@login_required(login_url="/login/")
@vezetoi
def Foglalkozas_tipusList(request):
    context = {}
    foglalkozas_tipusok = Foglalkozas_tipus.objects.all()
    myFilter = Foglalkozas_tipusFilter(request.GET, queryset=foglalkozas_tipusok)
    foglalkozas_tipusok = myFilter.qs
    context['foglalkozas_tipusok'] = foglalkozas_tipusok
    context['myFilter'] = myFilter
    return render(request, 'foglalkozas_tipus_list.html', context)

@login_required(login_url='login')
@vezetoi
def Foglalkozas_tipusUpdate(request, pk):
    foglalkozas_tipus = Foglalkozas_tipus.objects.get(id=pk)
    form = Foglalkozas_tipusForm(instance=foglalkozas_tipus)
    context ={}
    if request.method == 'POST':
        form = Foglalkozas_tipusForm(request.POST, instance=foglalkozas_tipus)
        if form.is_valid():
            form.save()
            return redirect('foglalkozas_tipus_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Foglalkozás típus módósító űrlap'
    return render(request, 'create_form.html', context)



#------------------------------------------- Csoport views--------------------------------------------------------------------------------------------

@login_required(login_url='login')
@vezetoi
def CsoportCreate(request):
    form = CsoportForm()
    csoport = Csoport.objects.none()
    guru_formset = GurutoCsoportFormSet()
    gyerek_formset = GyerektoCsoportFormSet()
    context ={}
    
    
    if request.method == 'POST':
        form = CsoportForm(request.POST)
        guru_formset = GurutoCsoportFormSet(request.POST)
        gyerek_formset = GyerektoCsoportFormSet(request.POST)

        if form.is_valid():
            created_csoport = form.save(commit=False)
            guru_formset = GurutoCsoportFormSet(request.POST, instance=created_csoport )
            gyerek_formset = GyerektoCsoportFormSet(request.POST, instance=created_csoport)
            if gyerek_formset.is_valid() and guru_formset.is_valid():
                
                created_csoport.save()
                gyerek_formset.save()
                guru_formset.save()
                return redirect('terapia_lista')
                

                
    context = {'form': form, 'guru_formset':guru_formset, 'gyerek_formset':gyerek_formset }
    return render(request, 'terapia.html', context)

@login_required(login_url="/login/")
@emusok
def CsoportList(request):
    context = {}
    user =request.user
    group = user.groups.all()[0].name
 

    if group == 'terapeuta':
        guru = user.guru
        qs = GurutoCsoport.objects.values_list('csoport', flat=True).filter(guru=guru)
        csoportok = Csoport.objects.filter(id__in=qs)
        myFilter = CsoportFilter(queryset=csoportok)
        csoportok = myFilter.qs

    else:
        csoportok = Csoport.objects.all()
        myFilter = CsoportFilter(queryset=csoportok)
        csoportok = myFilter.qs

    context['csoportok'] = csoportok
    
    return render(request, 'terapia_list.html', context)

@login_required(login_url="/login/")
@emusok
def CsoportListExport(request):
    user = request.user
    group = user.groups.all()[0].name

    if group == 'terapeuta':
        guru = user.guru
        qs = GurutoCsoport.objects.values_list('csoport', flat=True).filter(guru=guru)
        csoportok = Csoport.objects.filter(id__in=qs).filter(aktiv=True)
        myFilter = CsoportFilter(request.GET, queryset=csoportok)
        csoportok = myFilter.qs
        wb_filename=guru.nev+"_terápiái.xls"
        sheet_name=guru.nev+" terápiái"

    else:
        csoportok = Csoport.objects.all()
        myFilter = CsoportFilter(request.GET, queryset=csoportok)
        csoportok = myFilter.qs
        wb_filename="terápiák.xls"
        sheet_name="terápiák"
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=terapiak.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(sheet_name)
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['id', 'Név', 'Típus', 'Helyszin', 'Aktív']

    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style.font.bold = False
    row_num = 0

    for item in csoportok:
        row_num += 1
        ws.write(row_num, 0, str(getattr(item, 'id')), font_style)
        ws.write(row_num, 1, str(getattr(item, 'nev')), font_style)
        ws.write(row_num, 2, str(getattr(item, 'foglalkozas_tipus')), font_style)
        ws.write(row_num, 3, str(getattr(item, 'helyszin')), font_style)
        ws.write(row_num, 4, str(getattr(item, 'aktiv')), font_style)
        
    
    wb.save(response)
    return response

@login_required(login_url="/login/")
@emusok
def CsoportUpdate(request, pk): # https://stackoverflow.com/questions/27876644/django-class-based-createview-and-updateview-with-multiple-inline-formsets

    csoport = Csoport.objects.get(id=pk)  # if this is an edit form, replace the author instance with the existing one
    form = CsoportForm(instance=csoport)
    
    gyerek_formset = GyerektoCsoportFormSet(instance=csoport)
    guru_formset = GurutoCsoportFormSet(instance=csoport)
    orak = Ora.objects.filter(csoport=csoport)

    group = request.user.groups.all()[0].name
    if group == 'terapeuta':        
        guru = request.user.guru
        qs = GurutoCsoport.objects.values_list('guru', flat=True).filter(csoport=csoport)
        if guru.id not in qs:
            return HttpResponse('Ez nem a te Terápiád, ne huncutkodj!')


    if request.method == "POST":
        form = CsoportForm(request.POST)

        if pk: 
            form = CsoportForm(request.POST, instance=csoport)
            orak = Ora.objects.filter(csoport=csoport)

        
        gyerek_formset = GyerektoCsoportFormSet(request.POST)
        guru_formset = GurutoCsoportFormSet(request.POST)

        if form.is_valid():
            created_csoport = form.save(commit=False)
            gyerek_formset = GyerektoCsoportFormSet(request.POST, instance=created_csoport)
            guru_formset = GurutoCsoportFormSet(request.POST, instance=created_csoport)

            if gyerek_formset.is_valid() and guru_formset.is_valid():
                created_csoport.save()
                gyerek_formset.save()
                guru_formset.save()

            # Jelenlétik  amelyek a gomb megnyomásakor már le vannak generálva, újra generálódnak
                terapia=csoport
                mai_nap=datetime.datetime.now()
                orak=Ora.objects.filter(csoport=terapia).filter(kezdes__gt=mai_nap)
                for ora in orak:        
                    if not ora.megtartott or ora.lezart:
                        Jelenleti.objects.filter(ora=ora).delete()
                        csoport = ora.csoport
                        gyerekektocsoport = GyerektoCsoport.objects.filter(csoport=csoport)
                        for gyerektocsoport in gyerekektocsoport:
                                    gyerek = gyerektocsoport.gyerek
                                    jelenlet = Jelenleti(gyerek=gyerek, ora=ora, status=0)
                                    jelenlet.save()
    context = {
        'form': form,
        'guru_formset':guru_formset, 
        'gyerek_formset':gyerek_formset,
        'csoport': csoport,
        "orak" : orak,
    }
        
    
    return render(request, 'terapia_mod.html', context)




@login_required(login_url="/login/")
def OraCreate(request):
    
    user = request.user
    form = OraCreateForm()
    context ={}
    csoportok = Csoport.objects.none()
    group = user.groups.all()[0].name

    if group == 'terapeuta':
        guru = user.guru
        qs = GurutoCsoport.objects.values_list('csoport', flat=True).filter(guru=guru)
        form.fields['csoport'].queryset = Csoport.objects.filter(id__in=qs).filter(aktiv=True)
    else:
        csoportok = Csoport.objects.all()
      
    
    if request.method == 'POST':
        form = OraCreateForm(request.POST)
        if group == 'terapeuta':
            form.fields['csoport'].queryset = Csoport.objects.filter(id__in=qs).filter(aktiv=True)
        
        if form.is_valid():
            ora = form.save(commit=False)
            ora.save()
            #jelenleti ivek generalasa
            JelenletiGen(ora.id)


            return redirect('ora_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Óra felvétel űrlap'
    return render(request, 'oracreate.html', context)


@login_required(login_url="/login/")
def OraIsmetles(request, pk):
    ora = Ora.objects.get(id=pk)
    
    
    group = request.user.groups.all()[0].name
    if group == 'terapeuta':        
        guru = request.user.guru
        qs = GurutoCsoport.objects.values_list('guru', flat=True).filter(csoport=ora.csoport)
        if guru.id not in qs:
            return HttpResponse('Ez nem a te órád, ne huncutkodj!')
  
    form = OraismetlesForm()
    ora = Ora.objects.get(id=pk)
    startok = datetime.datetime
    endek = datetime.datetime
    elsokezdes = datetime.datetime
    start = ora.kezdes
    end = ora.befejezes
    csoport=ora.csoport

    napok = []
    if request.method == 'POST':
        form = OraismetlesForm(request.POST)
        if form.is_valid():
            hetfo = form.cleaned_data['hetfo']
            if hetfo :
                napok.append(0)
            kedd = form.cleaned_data['kedd']
            if kedd :
                napok.append(1)
            szerda = form.cleaned_data['szerda']
            if szerda :
                napok.append(2)
            csutortok = form.cleaned_data['csutortok']
            if csutortok :
                napok.append(3)
            pentek = form.cleaned_data['pentek']
            if pentek :
                napok.append(4)
            szombat = form.cleaned_data['szombat']
            if szombat :
                napok.append(5)
            
                # a napok változó eredménye, ha be volt jelölve a hétfő és a szerda: (0,2)
            periodus = form.cleaned_data['periodus']
            idohosz = form.cleaned_data['idohosz']
            idohosz = idohosz + 1# korrekciós
            if periodus == '1':
                startok = rrule(DAILY, count=idohosz, byweekday=napok, dtstart=start)
                endek = rrule(DAILY, count=idohosz, byweekday=napok, dtstart=end)
            elif periodus == '2':
                startok = rrule(WEEKLY, count=idohosz, byweekday=napok, dtstart=start)
                endek = rrule(WEEKLY, count=idohosz, byweekday=napok, dtstart=end)
            elif periodus == '3':
                startok = rrule(MONTHLY, count=idohosz, dtstart=start)
                endek = rrule(MONTHLY, count=idohosz, dtstart=end)
            
            idopontok_szama = len(list(startok))
            elsokezdes=startok[0]
            for i in range(1,idopontok_szama):
                kezdes=startok[i]
                befejezes=endek[i]
                newora = Ora.objects.none()
                newora = Ora(kezdes=kezdes, befejezes=befejezes, csoport=csoport, feljegyzes='none', megtartott=False)
                newora.save()
                newora.id
                JelenletiGen(newora.id)
                             

            return redirect ('ora_lista')
    
    context = {'form': form, 'ora':ora}
    context['Form_neve'] = 'Ora felvétel űrlap'
    return render(request, 'ismetles.html', context)
    
@login_required(login_url="/login/")
def OraListFull(request):
     
    context = {}
    group = request.user.groups.all()[0].name
    orak = Ora.objects.none()
    terapeuta_nev=""

    if group == 'vezeto':
        orak = Ora.objects.all()

    if group == 'terapeuta':
        guru = request.user.guru
        terapeuta_nev=guru.nev
        gurucsoportok = GurutoCsoport.objects.filter(guru=guru)
        for gurucsoport in gurucsoportok:
            csoport = gurucsoport.csoport
            orak |= csoport.ora_set.all()

    context['cim'] = "Összes bejegyzett óra"
    context['orak'] = orak
    context['terapeuta_nev'] = terapeuta_nev
    
    return render(request, 'ora_list.html', context)

@login_required(login_url="/login/")
def OraListEgy(request):
     
    context = {}
    group = request.user.groups.all()[0].name
    orak = Ora.objects.none()
    terapeuta_nev=""

    if group == 'vezeto':
        orak = Ora.objects.all().filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month)

    if group == 'terapeuta':
        guru = request.user.guru
        terapeuta_nev=guru.nev
        gurucsoportok = GurutoCsoport.objects.filter(guru=guru)
        for gurucsoport in gurucsoportok:
            csoport = gurucsoport.csoport
            orak |= csoport.ora_set.all()
        orak = orak.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month)
        
        
    context['ev'] = datetime.datetime.now().year
    context['honap'] = datetime.datetime.now().month
    context['cim'] = "havi órák"
    context['orak'] = orak
    context['terapeuta_nev'] = terapeuta_nev
    
    return render(request, 'ora_list.html', context)

@login_required(login_url="/login/")
def OraListEgyMeg(request):
     
    context = {}
    group = request.user.groups.all()[0].name
    orak = Ora.objects.none()
    terapeuta_nev=""

    if group == 'vezeto':
        orak = Ora.objects.all().filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month, megtartott=True)

    if group == 'terapeuta':
        guru = request.user.guru
        terapeuta_nev=guru.nev
        gurucsoportok = GurutoCsoport.objects.filter(guru=guru)
        for gurucsoport in gurucsoportok:
            csoport = gurucsoport.csoport
            orak |= csoport.ora_set.all()
        orak = orak.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month, megtartott=True)

    context['cim'] = "Aktuális megtartott havi orák"
    context['orak'] = orak
    context['terapeuta_nev'] = terapeuta_nev
    
    return render(request, 'ora_list.html', context)

@login_required(login_url="/login/")
def OraListEgyNem(request):
     
    context = {}
    group = request.user.groups.all()[0].name
    orak = Ora.objects.none()
    terapeuta_nev=""

    if group == 'vezeto':
        orak = Ora.objects.all().filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month, megtartott=False)

    if group == 'terapeuta':
        guru = request.user.guru
        terapeuta_nev=guru.nev
        gurucsoportok = GurutoCsoport.objects.filter(guru=guru)
        for gurucsoport in gurucsoportok:
            csoport = gurucsoport.csoport
            orak |= csoport.ora_set.all()
        orak = orak.filter(kezdes__year=datetime.datetime.now().year, kezdes__month=datetime.datetime.now().month, megtartott=False)

    context['cim'] = "Aktuális még meg nem tartott havi orák"
    context['orak'] = orak
    context['terapeuta_nev'] = terapeuta_nev
    
    return render(request, 'ora_list.html', context)




@login_required(login_url="/login/")
def OraListHarom(request):
     
    context = {}
    group = request.user.groups.all()[0].name
    orak = Ora.objects.none()
    terapeuta_nev=""

    ma=datetime.date.today()
    date_tmp = ma+relativedelta(months=-1)
    indulo=date_tmp+relativedelta(day=1)
    date_tmp = ma+relativedelta(months=+1)
    vege = date_tmp+relativedelta(day=31)

    if group == 'vezeto':
        orak = Ora.objects.all().filter(kezdes__gte=indulo, kezdes__lte=vege)

    if group == 'terapeuta':
        guru = request.user.guru
        terapeuta_nev=guru.nev
        
        gurucsoportok = GurutoCsoport.objects.filter(guru=guru)
        for gurucsoport in gurucsoportok:
            csoport = gurucsoport.csoport
            orak |= csoport.ora_set.all()
        orak = orak.filter(kezdes__gte=indulo, kezdes__lte=vege)

    context['orak'] = orak
    context['terapeuta_nev'] = terapeuta_nev
    context['cim'] = "A múlt hónaptól a következő hónap végéig bejegyzett órák"
    
    return render(request, 'ora_list.html', context)




@login_required(login_url="/login/")
def OraUpdate(request, pk):
    ora = Ora.objects.get(id=pk)

    group = request.user.groups.all()[0].name
    if group == 'terapeuta':        
        guru = request.user.guru
        qs = GurutoCsoport.objects.values_list('guru', flat=True).filter(csoport=ora.csoport)
        if guru.id not in qs:
            return HttpResponse('Ez nem a te órád, ne huncutkodj!')

    form = OraModForm(instance=ora)
    context = {}
    if request.method == 'POST':
        form = OraModForm(request.POST, instance=ora)
        if form.is_valid():
            form.save()
            return redirect('ora_lista')
    context = {'form':form, 'ora':ora}
    return render (request, 'oramod.html', context)



@login_required(login_url="/login/")
def OraTorles (request, pk):
    msg = 'Az óra meg lett tartva ezért nem törölhető!'
    ora = Ora.objects.get(id=pk)
    group = request.user.groups.all()[0].name
    if group == 'terapeuta':        
        guru = request.user.guru
        qs = GurutoCsoport.objects.values_list('guru', flat=True).filter(csoport=ora.csoport)
        if guru.id not in qs:
            return HttpResponse('Ez nem a te órád, ne huncutkodj!')

    jelenletik = Jelenleti.objects.filter(ora=ora)

    if request.method == "POST":
        if ora.megtartott :
            return HttpResponse(msg)
        else:         
            for jelenleti in jelenletik:
                jelenleti.delete()
            ora.delete()
            return redirect('ora_lista')
        
    link = "ora_torles"
    context = {'objektum':ora, 'link':link}
    return render(request, 'delete.html', context)


@login_required(login_url="/login/")
def JelenletiView(request, pk):
    context ={}
    save_msg = ""
    jelenleti_msg = ""
    ora = Ora.objects.get(id=pk)
    
    csoport=ora.csoport
    gycsqs = GyerektoCsoport.objects.values_list('gyerek', flat=True).filter(csoport=csoport)
    gyerekek = Gyerek.objects.filter(id__in=gycsqs)
    jelenletik=Jelenleti.objects.filter(ora=pk) # lekérdezi az órákhoz tartozó jelenléti íveket

    if not jelenletik: # ha még az órához nincs jelenléti adat, létrehozza azt
        for gyerek in gyerekek:
            
            jelenlet = Jelenleti(gyerek=gyerek, ora=ora, status=0)
            jelenlet.save()
            jelenleti_msg="Az órához tartozó jelenléti űrlap most generálódott."

    group = request.user.groups.all()[0].name
    if group == 'terapeuta':        
        guru = request.user.guru
        qs = GurutoCsoport.objects.values_list('guru', flat=True).filter(csoport=ora.csoport)
        if guru.id not in qs:
            return HttpResponse('Ez nem a te órád, ne huncutkodj!')
            
    JelenletiFormSet = inlineformset_factory(Ora, Jelenleti, form=JelenletiForm2, fields=('gyerek', 'status' ), extra=0)
    formset = JelenletiFormSet(instance=ora, form_kwargs={'pk': pk})
    
    
    guruk = GurutoCsoport.objects.filter(csoport=csoport)
    
    

    if request.method == 'POST':
        formset = JelenletiFormSet(request.POST, instance=ora, form_kwargs={'pk': pk})
        if formset.is_valid():
            # ha a jelenléti formsetben valamelyik status mezo nem 0, akkor az órát aktívra kell állítani
            st = 0
            for f in formset:
                cd = f.cleaned_data
                status = cd.get('status')
                st = st + status
                if status != 0:
                    ora.megtartott=True
                else:
                    ora.megtartott=False
                
                    
            ora.save(update_fields=['megtartott'])

            formset.save()
            save_msg = "A jelenléti ív sikeresen elmentve"

            context = {'formset':formset, 'ora':ora, 'csoport':csoport, 'guruk':guruk, 'msg':save_msg}
            return render(request, 'jelenleti.html', context)
        
    
    context = {'formset':formset, 'ora':ora, 'csoport':csoport, 'guruk':guruk, 'msg':jelenleti_msg}
    return render(request, 'jelenleti.html', context)




def Jelenleti_regenerate_View(request, pk):
    context ={}    
    ora = Ora.objects.get(id=pk)
    jelenletik=Jelenleti.objects.filter(ora=pk) # lekérdezi az órákhoz tartozó jelenléti íveket
    group = request.user.groups.all()[0].name

    if ora.lezart :
            return HttpResponse("Az óra és a hozzá tartozó jelenléti ív már nem módosítható")

    if not jelenletik:
        # ha még az órához nincs jelenléti adat, hibára figyelmeztet
        return HttpResponse('Nem létező jelenléti ívet nem lehet újragenerálni!')
        
    if group == 'terapeuta':        
        guru = request.user.guru
        qs = GurutoCsoport.objects.values_list('guru', flat=True).filter(csoport=ora.csoport)
        if guru.id not in qs:
            return HttpResponse('Ez nem a te órád, ne huncutkodj!')
    
    if request.method == "POST":
    
        for jelenleti in jelenletik:
            jelenleti.delete()
        csoport = ora.csoport
        gyerekektocsoport = GyerektoCsoport.objects.filter(csoport=csoport)
        for gyerektocsoport in gyerekektocsoport:
            gyerek = gyerektocsoport.gyerek
            jelenlet = Jelenleti(gyerek=gyerek, ora=ora, status=0)
            jelenlet.save()
        
        return redirect('jelenleti', ora.id)
           
    
    link = "jelenleti_regen"
    context = {'objektum':ora, 'link':link}
    return render(request, 'jelenleti_regen.html', context)


@login_required(login_url="/login/")
@vezetoi
def BefizetesView(request):
    form = BefizetesForm()
    context ={}
    if request.method == 'POST':
        form = BefizetesForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('gyerek_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Befizetés űrlap'
    return render(request, 'create_form.html', context)

@login_required(login_url="/login/")
@vezetoi
def BefizetesList(request):
    context = {}
    befizetesek = Befizetes.objects.all()
    myFilter = BefizetesFilter(request.GET, queryset=befizetesek)
    befizetesek = myFilter.qs
    context['befizetesek'] = befizetesek
    context['myFilter'] = myFilter
    return render(request, 'befizetes_list.html', context)


@login_required(login_url='login')
@vezetoi
def BefizetesUpdate(request, pk):
    befizetes=Befizetes.objects.get(id=pk)
    form = BefizetesForm(instance=befizetes)
    context ={}
    if request.method == 'POST':
        form = BefizetesForm(request.POST, instance=befizetes)
        if form.is_valid():
            form.save()
            return redirect('befizetes_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Befizetés módosító űrlap'
    return render(request, 'create_form.html', context)



#def OraJovahagyo(request, pk):


# FullCallendar használata:
# https://stackoverflow.com/questions/39902405/fullcalendar-in-django
 
@login_required(login_url="/login/")
def naptar(request):

    
    user = request.user
    group = request.user.groups.all()[0].name
    ma=datetime.date.today()
    date_tmp = ma+relativedelta(months=-2)
    indulo=date_tmp+relativedelta(day=1)
    date_tmp = ma+relativedelta(months=+2)
    vege = date_tmp+relativedelta(day=31)

    
    if group == 'vezeto':
        orak = Ora.objects.all().filter(kezdes__gte=indulo, kezdes__lte=vege)
        egyediek = Egyedi.objects.filter(idopont__gte=indulo, idopont__lte=vege)
     
    if group == 'terapeuta':
        guru = user.guru
        qs = GurutoCsoport.objects.values_list('csoport', flat=True).filter(guru=guru)
        orak = Ora.objects.filter(csoport__in=qs)
        orak = orak.filter(kezdes__gte=indulo, kezdes__lte=vege)       
        egyediek = Egyedi.objects.none()

    context = {"orak":orak, "egyediek":egyediek}
    return render(request,'calendar.html',context)

#-------------------------   Adminisztráció ---------------------------------------------------------------------

@login_required(login_url="/login/")
@vezetoi
def register_user(request):

    msg     = None
    success = False

    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get("username")
            raw_password = form.cleaned_data.get("password1")
            user = authenticate(username=username, password=raw_password)

            msg     = 'A felhasználó elkészült! <a href="/login">login</a>.'
            success = True
            
            return redirect("/login/")

        else:
            msg = 'Az űrlap kitöltése nem megfelelő!'    
    else:
        form = SignUpForm()

    return render(request, "accounts/register.html", {"form": form, "msg" : msg, "success" : success })


@login_required(login_url="/login/")
@vezetoi
def felhasznalok(request):

    context = {}
    felhasznalok = User.objects.all()
    
    context['felhasznalok'] = felhasznalok
    
    return render(request, 'felhasznalok_list.html', context)

@login_required(login_url="/login/")
@vezetoi
def user_mod(request, pk):
    user = User.objects.get(id=pk)
    form = UserModForm(instance=user)
    context ={}
    if request.method == 'POST':
        form = UserModForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('felhasznalok')
                
    context = {'form': form, 'user':user}
    context['Form_neve'] = 'Felhasználó módosító űrlap'
    return render(request, 'user_mod.html', context)


@login_required(login_url="/login/")
def jelszocsere(request):
    context ={}
    user = request.user
    if request.method == 'POST':
        
        form = JelszocsereForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, _('A jelszava megváltozott'))
            return redirect('home')
        else:
            messages.error(request, _('Kérem javítsa a hibát!'))
    else:
        form = JelszocsereForm(request.user)
    
    context = {'form': form, 'user':user}
    return render(request, 'accounts/passchange.html', context)


@login_required(login_url="/login/")
@vezetoi
def passwchange(request, pk):
    context ={}
    user = User.objects.get(id=pk)

    if request.method == 'POST':
        form = PasswChangeForm(user, request.POST)

        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, _('A jelszó megváltozott'))
            return redirect('home')

        else:
            messages.error(request, _('Kérem javítsa a hibát!'))
    else:
        form = PasswChangeForm(user)
    
    context = {'form': form, 'user':user}
    return render(request, 'accounts/passwchange.html', context)




@login_required(login_url="/login/")
@vezetoi
def EgyediCreate(request):
    
    form = EgyediForm()
    context ={}
    if request.method == 'POST':
        form = EgyediForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('egyedi_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Egyedi alkalom felvétel űrlap'
    return render(request, 'create_form.html', context)


@login_required(login_url="/login/")
@vezetoi
def EgyediList(request):
    context = {}
    egyediek = Egyedi.objects.all()
    myFilter = EgyediFilter(request.GET, queryset=egyediek)
    egyediek = myFilter.qs
    context['egyediek'] = egyediek
    context['myFilter'] = myFilter
    return render(request, 'egyedi_list.html', context)


@login_required(login_url="/login/")
@vezetoi
def EgyediUpdate(request, pk):
    egyedi = Egyedi.objects.get(id=pk)
    form = EgyediForm(instance=egyedi)
    context ={}
    if request.method == 'POST':
        form = EgyediForm(request.POST, instance=egyedi)
        if form.is_valid():
            form.save()
            return redirect('egyedi_lista')

    context = {'form': form}
    context['Form_neve'] = 'Egyedi alkalom módosítás űrlap'
    return render(request, 'create_form.html', context)


@login_required(login_url="/login/")
@vezetoi
def EgyediTorles (request, pk):
    msg = 'Az alkalmat nem lehet törölni, mert már megtartott!'
    egyedi = Egyedi.objects.get(id=pk)
    if request.method == "POST":
        if egyedi.megtartva :
            return HttpResponse(msg)
        else:         
            egyedi.delete()
            return redirect('egyedi_lista')
        

    link = "egyedi_torles"
    context = {'objektum':egyedi, 'link':link}
    return render(request, 'delete.html', context)


@login_required(login_url="/login/")
@vezetoi
def NapiuziCreate(request):
    
    form = NapiuziForm()
    context ={}
    if request.method == 'POST':
        
        form = NapiuziForm(request.POST)
        if form.is_valid():
            napiuzi = form.save(commit=False)
            napiuzi.rogzito = request.user
            napiuzi.save()
            return redirect('napiuzi_lista')
                
    context = {'form': form}
    context['Form_neve'] = 'Napi üzenet létrehozása'
    return render(request, 'create_form.html', context)


@login_required(login_url="/login/")
@vezetoi
def NapiuziList(request):
    context = {}
    napiuzik = Napiuzi.objects.all()
    myFilter = NapiuziFilter(request.GET, queryset=napiuzik)
    napiuzik = myFilter.qs
    context['napiuzik'] = napiuzik
    context['myFilter'] = myFilter
    return render(request, 'napiuzi_list.html', context)


@login_required(login_url="/login/")
@vezetoi
def NapiuziUpdate(request, pk):
    napiuzi = Napiuzi.objects.get(id=pk)
    form = NapiuziForm(instance=napiuzi)
    context ={}
    if request.method == 'POST':
        form = NapiuziForm(request.POST, instance=napiuzi)
        if form.is_valid():
            napiuzi = form.save(commit=False)
            napiuzi.rogzito = request.user
            napiuzi.save()
            return redirect('napiuzi_lista')

    context = {'form': form}
    context['Form_neve'] = 'Napi üzenet módosítás űrlap'
    return render(request, 'create_form.html', context)

@login_required(login_url="/login/")
@vezetoi
def NapiuziTorles (request, pk):
    objektum = Napiuzi.objects.get(id=pk)
    if request.method == "POST":
        objektum.delete()
        return redirect('/')
        
    link = "napiuzi_torles"
    context = {'objektum':objektum, 'link':link}
    return render(request, 'delete.html', context)



    
@login_required(login_url="/login/")
@vezetoi
def EgyenlegekView(request):

    context = {}
    tablazat=[]
    szulok=Szulo.objects.filter(szerzodo=True).filter(aktiv=True).order_by('nev')
    for szulo in szulok:
        sor = []
        oradijak = 0
        csalad_koltsege=0
        
        osszeg = Befizetes.objects.filter(szulo=szulo).aggregate(Sum('osszeg'))['osszeg__sum']# a szülő által összesen befizetett pénz
        
        sor.append(szulo.id)
        sor.append(szulo.nev)
        sor.append(osszeg)
        csalad = Csalad.objects.filter(szulo=szulo)
        

        for kapcsolat in csalad:
            gyerek=kapcsolat.gyerek
            
            sor.append(gyerek.kod)
            jelenletik = Jelenleti.objects.filter(gyerek=gyerek).filter( Q(status=1) | Q(status=2) )# jelen volt, vagy 100%-ban lemondott órák Jelenléti objektumok
            sor.append(jelenletik.count())
            for jelenleti in jelenletik:
                oradijak = oradijak + jelenleti.ora.csoport.foglalkozas_tipus.ar
            sor.append(oradijak)
        
        tablazat.append(sor)
    
    context = {'tablazat':tablazat}
     
    return render(request, 'egyenlegek.html', context)

@login_required(login_url="/login/")
@vezetoi
def Osszesített_elszamolas(request):
    context = {}
    lista=[]
    ev=datetime.datetime.now().year
    elozo_honap=datetime.datetime.now().month-1
    
    gyerekek = Gyerek.objects.filter(aktiv=True)
    for gyerek in gyerekek:
        oradijak = 0
        gykod=gyerek.kod
        gyid=gyerek.id
        qs = Jelenleti.objects.values_list('ora', flat=True).filter(gyerek=gyerek).filter( Q(status=1) | Q(status=3) )# jelen volt, vagy 100%-ban lemondott órák Jelenléti objektumok
        orak = Ora.objects.filter(id__in=qs).filter(kezdes__year=ev, kezdes__month=elozo_honap)
        for ora in orak:
            oradijak = oradijak + ora.csoport.foglalkozas_tipus.ar

        
        lista.append({'gyerek_id':gyid, 'gyerek_kod':gykod, 'oradijak':oradijak})


    context['lista'] = lista
    
    return render(request, 'osszesitett_elszamolas.html', context)


@login_required(login_url="/login/")
@vezetoi
def export_havi_osszesites(request, pk):

    gyerek = Gyerek.objects.get(id=pk)
    ev=datetime.datetime.now().year
    elozo_honap=datetime.datetime.now().month-1
    osszesen=0

    #workbook = load_workbook('media/havi_elszamolas_2.xlsx')
    workbook = load_workbook('/home/zmogrlnc/emu/media/havi_elszamolas_2.xlsx')
    
    worksheet = workbook.active
    worksheet.title = gyerek.nev
    

    qs = Jelenleti.objects.values_list('ora', flat=True).filter(gyerek=gyerek)
    orak = Ora.objects.filter(id__in=qs).filter(kezdes__year=ev, kezdes__month=elozo_honap)
    fejlec=gyerek.nev+" terápiái a "+str(ev)+". év "+str(elozo_honap)+". hónapban"
    worksheet.cell(row=1, column=1).value=fejlec
    sor=4
    
    for ora in orak:
        worksheet.cell(row=sor, column=1).value=ora.csoport.nev
        worksheet.cell(row=sor, column=2).value=ora.csoport.foglalkozas_tipus.nev
        worksheet.cell(row=sor, column=3).value=ora.kezdes.date()
        allapot=Jelenleti.objects.filter(gyerek=gyerek).filter(ora=ora)[0].status
        if allapot == 0:
            worksheet.cell(row=sor, column=4).value="jelöletlen"
            worksheet.cell(row=sor, column=5).value=0
        elif allapot == 1:
            worksheet.cell(row=sor, column=4).value="megjelent"
            worksheet.cell(row=sor, column=5).value=ora.csoport.foglalkozas_tipus.ar
            osszesen = osszesen + ora.csoport.foglalkozas_tipus.ar
        elif allapot == 2:
            worksheet.cell(row=sor, column=4).value="0%-os lemondás"
            worksheet.cell(row=sor, column=5).value=0
        elif allapot == 3:
            worksheet.cell(row=sor, column=4).value="100%-os lemondás"
            worksheet.cell(row=sor, column=5).value=ora.csoport.foglalkozas_tipus.ar
            osszesen = osszesen + ora.csoport.foglalkozas_tipus.ar
        sor = sor + 1

    worksheet.cell(row=sor+1, column=1).value="Összesen:"
    worksheet.cell(row=sor+1, column=5).value=osszesen
    xlsfilename=gyerek.nev+"_"+str(ev)+"_"+str(elozo_honap)+".xlsx"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(content=output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={xlsfilename}'
    return response
    '''
    response = HttpResponse(content=save_virtual_workbook(workbook))
    response['Content-Disposition'] = 'attachment; filename='+xlsfilename
    return response

    '''
    
@login_required(login_url="/login/")
@vezetoi
def export_havi_osszesites_honap(request, pk):

    ev=pk[:4]
    honap=pk[4:6]
    gyerekek = Gyerek.objects.filter(aktiv=True).order_by('nev')
    
    #workbook = load_workbook('static/havi_elszamolas_all.xlsx')
    workbook = load_workbook('/home/zmogrlnc/emu/media/havi_elszamolas_all.xlsx')
    fejlec="A lehallgatott terápiák a "+str(ev)+". év "+str(honap)+". hónapban"
    
    worksheet = workbook.active
    worksheet.title = str(ev)+". év "+str(honap)+". hónap"
    worksheet.cell(row=1, column=1).value=fejlec
    sor=4
    kintlevoseg = 0

    for gyerek in gyerekek:

        osszesen=0
        qs = Jelenleti.objects.values_list('ora', flat=True).filter(gyerek=gyerek)
        orak = Ora.objects.filter(id__in=qs).filter(kezdes__year=ev, kezdes__month=honap)
        sor = sor +2
    
        for ora in orak:
            worksheet.cell(row=sor, column=1).value=gyerek.nev
            worksheet.cell(row=sor, column=2).value=ora.csoport.nev
            worksheet.cell(row=sor, column=3).value=ora.csoport.foglalkozas_tipus.nev
            worksheet.cell(row=sor, column=4).value=ora.csoport.foglalkozas_tipus.ar
            worksheet.cell(row=sor, column=5).value=ora.kezdes.date()
            allapot=Jelenleti.objects.filter(gyerek=gyerek).filter(ora=ora)[0].status
            
            if allapot == 0:
                worksheet.cell(row=sor, column=6).value="jelöletlen"
                worksheet.cell(row=sor, column=7).value=0
            elif allapot == 1:
                worksheet.cell(row=sor, column=6).value="megjelent"
                worksheet.cell(row=sor, column=7).value=ora.csoport.foglalkozas_tipus.ar
                osszesen = osszesen + ora.csoport.foglalkozas_tipus.ar
            elif allapot == 2:
                worksheet.cell(row=sor, column=6).value="0%-os lemondás"
                worksheet.cell(row=sor, column=7).value=0
            elif allapot == 3:
                worksheet.cell(row=sor, column=6).value="100%-os lemondás"
                worksheet.cell(row=sor, column=7).value=ora.csoport.foglalkozas_tipus.ar
                osszesen = osszesen + ora.csoport.foglalkozas_tipus.ar
            sor = sor + 1

        
        worksheet.cell(row=sor, column=8).value=osszesen
        kintlevoseg = kintlevoseg + osszesen

    worksheet.cell(row=sor+1, column=8).value=kintlevoseg
    xlsfilename="Összesítö_"+str(ev)+"_"+str(honap)+".xlsx"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(content=output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={xlsfilename}'
    return response
    '''
    response = HttpResponse(content=save_virtual_workbook(workbook))
    #response = HttpResponse(content=save_virtual_workbook(workbook), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+xlsfilename
    return response

    '''
    
@login_required(login_url="/login/")
@vezetoi
def osszesito(request):
    form = Havi_osszesites()
    if request.method == 'POST':
        form = Havi_osszesites(request.POST)
        if form.is_valid():
            ev = form.cleaned_data.get("ev")
            honap = form.cleaned_data.get("honap")
            evhonap = ev+honap
            return redirect ('export_havi_osszesites_honap', pk=evhonap)
    return render(request, 'osszesito.html',{'form': form})


    
         # lekérdezi az órákhoz tartozó je
